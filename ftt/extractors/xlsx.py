from __future__ import annotations

from pathlib import Path
from typing import List

from ftt.extractors.base import ExtractedContent, ImageRef
from ftt.logging_utils import FileLogger
from ftt.render import convert_office_to_pdf, render_pdf_pages


def _office_enabled(mode: str) -> bool:
    return mode.lower() != "false"


def _sheet_to_markdown(sheet) -> str:
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    rows: List[List[str]] = []
    for r in range(1, max_row + 1):
        row_values = []
        has_value = False
        for c in range(1, max_col + 1):
            value = sheet.cell(row=r, column=c).value
            text = "" if value is None else str(value)
            if text:
                has_value = True
            row_values.append(text)
        if has_value:
            rows.append(row_values)

    if not rows:
        return ""

    header = rows[0]
    body = rows[1:] if len(rows) > 1 else []
    header_line = "| " + " | ".join(header) + " |"
    separator = "| " + " | ".join("---" for _ in header) + " |"
    body_lines = ["| " + " | ".join(row) + " |" for row in body]
    return "\n".join([header_line, separator] + body_lines)


def extract_xlsx(
    path: Path,
    visual_mode: str,
    render_office_mode: str,
    render_dpi: int,
    render_max_pages: int,
    visuals_dir: Path,
    work_dir: Path,
    logger: FileLogger,
) -> ExtractedContent:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX extraction") from exc

    content = ExtractedContent()
    workbook = openpyxl.load_workbook(str(path), data_only=True)
    has_charts = False

    for sheet in workbook.worksheets:
        table_md = _sheet_to_markdown(sheet)
        if table_md:
            content.text_parts.append(f"[Sheet {sheet.title}]\n{table_md}")
        if getattr(sheet, "_charts", None):
            if len(sheet._charts) > 0:
                has_charts = True

    if visual_mode != "embedded" and _office_enabled(render_office_mode):
        if visual_mode == "full" or has_charts:
            pdf_path = convert_office_to_pdf(path, work_dir, logger)
            try:
                import pdfplumber
            except ImportError as exc:
                raise RuntimeError("pdfplumber is required for XLSX rendering") from exc
            with pdfplumber.open(str(pdf_path)) as pdf:
                total_pages = len(pdf.pages)
            pages = list(range(1, min(total_pages, render_max_pages) + 1))
            logger.info(f"Rendering {len(pages)} XLSX pages for visuals")
            image_paths = render_pdf_pages(pdf_path, visuals_dir, pages, render_dpi, logger)
            for image_path in image_paths:
                page_num = int(image_path.stem.split("_")[-1])
                content.images.append(ImageRef(path=image_path, label=f"sheet page {page_num}", source="xlsx"))

    return content
